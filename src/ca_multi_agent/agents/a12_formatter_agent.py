from typing import Dict, Any, List, Optional
import uuid
from datetime import datetime
import logging
import json
import pandas as pd
from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from .base_agent import BaseAgent

logger = logging.getLogger(__name__)

class ReportFormatterAgent(BaseAgent):
    def __init__(self):
        super().__init__("A12_Report_Formatter")
        self.supported_formats = ['pdf', 'excel', 'json', 'html']
        self.brand_templates = self._load_brand_templates()

    async def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        components = input_data.get('components', [])
        output_format = input_data.get('format', 'pdf')
        brand_theme = input_data.get('brand_theme', 'default')
        report_title = input_data.get('title', 'Report')
        
        if not components:
            raise ValueError("No components provided for formatting")
        
        if output_format not in self.supported_formats:
            raise ValueError(f"Unsupported format: {output_format}")
        
        # Generate the report
        report_content = await self._generate_report(components, output_format, brand_theme, report_title)
        
        # Save to storage (in real implementation)
        artifact_id = str(uuid.uuid4())
        file_path = f"/reports/{artifact_id}.{output_format}"
        
        return {
            'success': True,
            'artifact_id': artifact_id,
            'format': output_format,
            'file_path': file_path,
            'file_size': len(str(report_content)),
            'download_url': f"/api/v1/artifacts/{artifact_id}",
            'generated_at': datetime.now().isoformat(),
            'metadata': {
                'title': report_title,
                'component_count': len(components),
                'brand_theme': brand_theme
            }
        }

    async def _generate_report(self, components: List[Dict], format: str, brand_theme: str, title: str) -> Any:
        """Generate report in specified format"""
        if format == 'pdf':
            return await self._generate_pdf_report(components, brand_theme, title)
        elif format == 'excel':
            return await self._generate_excel_report(components, brand_theme, title)
        elif format == 'json':
            return await self._generate_json_report(components)
        elif format == 'html':
            return await self._generate_html_report(components, brand_theme, title)
        
        return None

    async def _generate_pdf_report(self, components: List[Dict], brand_theme: str, title: str) -> bytes:
        """Generate PDF report"""
        html_content = self._generate_html_content(components, brand_theme, title)
        
        try:
            pdf_bytes = HTML(string=html_content).write_pdf()
            return pdf_bytes
        except Exception as e:
            logger.error(f"PDF generation failed: {e}")
            # Fallback to simple PDF
            return self._generate_simple_pdf(components, title)

    async def _generate_excel_report(self, components: List[Dict], brand_theme: str, title: str) -> bytes:
        """Generate Excel report"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        
        # Apply brand styling
        self._apply_excel_styling(ws, brand_theme)
        
        # Add title
        ws['A1'] = title
        ws['A1'].font = Font(size=16, bold=True)
        
        row = 3
        for component in components:
            if component['type'] == 'table':
                self._add_excel_table(ws, component, row)
                row += len(component['data']) + 3
            elif component['type'] == 'chart':
                self._add_excel_chart(ws, component, row)
                row += 15
        
        # Save to bytes
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    async def _generate_json_report(self, components: List[Dict]) -> str:
        """Generate JSON report"""
        report_data = {
            'metadata': {
                'generated_at': datetime.now().isoformat(),
                'component_count': len(components)
            },
            'components': components
        }
        return json.dumps(report_data, indent=2)

    async def _generate_html_report(self, components: List[Dict], brand_theme: str, title: str) -> str:
        """Generate HTML report"""
        return self._generate_html_content(components, brand_theme, title)

    def _generate_html_content(self, components: List[Dict], brand_theme: str, title: str) -> str:
        """Generate HTML content for report"""
        template = self.brand_templates.get(brand_theme, self.brand_templates['default'])
        
        html_parts = [template['header'].replace('{{title}}', title)]
        
        for component in components:
            if component['type'] == 'table':
                html_parts.append(self._generate_html_table(component))
            elif component['type'] == 'chart':
                html_parts.append(self._generate_html_chart(component))
            elif component['type'] == 'text':
                html_parts.append(self._generate_html_text(component))
        
        html_parts.append(template['footer'])
        
        return '\n'.join(html_parts)

    def _generate_html_table(self, component: Dict) -> str:
        """Generate HTML table from component data"""
        df = pd.DataFrame(component['data'])
        return df.to_html(classes='table table-striped', index=False)

    def _generate_html_chart(self, component: Dict) -> str:
        """Generate HTML chart (using Chart.js or similar)"""
        # This would generate actual chart HTML
        return f"""<div class="chart">
            <h3>{component.get('title', 'Chart')}</h3>
            <p>Chart would be rendered here with data: {json.dumps(component['data'])[:100]}...</p>
        </div>"""

    def _load_brand_templates(self) -> Dict:
        """Load brand templates for different themes"""
        return {
            'default': {
                'header': """<!DOCTYPE html><html><head><meta charset="utf-8"><title>{{title}}</title>
                    <style>body{font-family:Arial,sans-serif;margin:40px;line-height:1.6}
                    table{border-collapse:collapse;width:100%;margin:20px 0}th,td{border:1px solid #ddd;padding:8px;text-align:left}
                    th{background-color:#f2f2f2}.chart{margin:30px 0}.section{margin:20px 0}
                    </style></head><body><h1>{{title}}</h1>""",
                'footer': """<footer><p>Generated on {date} by CA Multi-Agent System</p></footer></body></html>"""
                    .format(date=datetime.now().strftime('%Y-%m-%d %H:%M'))
            },
            'professional': {
                'header': """<!DOCTYPE html><html><head><meta charset="utf-8"><title>{{title}}</title>
                    <style>body{font-family:'Helvetica Neue',Arial,sans-serif;margin:60px;line-height:1.8;color:#333}
                    table{border-collapse:collapse;width:100%;margin:30px 0}th,td{border:1px solid #ccc;padding:12px;text-align:left}
                    th{background-color:#2c3e50;color:white}.chart{margin:40px 0}.section{margin:25px 0}
                    h1{color:#2c3e50;border-bottom:2px solid #3498db;padding-bottom:10px}
                    </style></head><body><h1>{{title}}</h1>""",
                'footer': """<footer style="margin-top:50px;padding-top:20px;border-top:1px solid #eee;color:#777">
                    <p>Generated on {date} | CA Multi-Agent System | Confidential</p></footer></body></html>"""
                    .format(date=datetime.now().strftime('%Y-%m-%d %H:%M'))
            }
        }

    def _apply_excel_styling(self, worksheet, brand_theme: str):
        """Apply brand-specific styling to Excel worksheet"""
        if brand_theme == 'professional':
            # Apply professional styling
            bold_font = Font(bold=True, color="FFFFFF")
            fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            for cell in worksheet[1]:
                cell.font = bold_font
                cell.fill = fill

    # Additional formatting methods would be implemented here...